import re
import os
from re import search
import xlrd
import numpy
import openpyxl
from openpyxl import Workbook
import tkinter
from tkinter import filedialog
import pandas as pd
import numpy as np

import validators
states = [ 'AK', 'AL', 'AR', 'AZ', 'CA', 'CO', 'CT', 'DC', 'DE', 'FL', 'GA',
           'HI', 'IA', 'ID', 'IL', 'IN', 'KS', 'KY', 'LA', 'MA', 'MD', 'ME',
           'MI', 'MN', 'MO', 'MS', 'MT', 'NC', 'ND', 'NE', 'NH', 'NJ', 'NM',
           'NV', 'NY', 'OH', 'OK', 'OR', 'PA', 'RI', 'SC', 'SD', 'TN', 'TX',
           'UT', 'VA', 'VT', 'WA', 'WI', 'WV', 'WY']

while True:
    state_cd = input("Enter the State CD (2 alphabets): ")
    if state_cd not in states :
        print("Not an appropriate value.")
        continue
    else:
        break

st_cd=' ' + state_cd + ' '
# load excel with its path
root = tkinter.Tk()
root.withdraw() #use to hide tkinter window

def browse_file():
   fname = filedialog.askopenfilename(filetypes = (("Template files", "*.xlsx"), ("All files", "*")))
   if len(fname) > 0:
       print("You chose: %s" % fname)
       return fname
   else: 
       print("No file chosen. Now exiting program")
       exit()

print('Please browse to the file.. File dialog window open now: ')
filename = browse_file()

# load county file with its path
root = tkinter.Tk()
root.withdraw() #use to hide tkinter window

def browse_file2():
   fname1 = filedialog.askopenfilename(filetypes = (("Template files", "*.xlsx"), ("All files", "*")))
   if len(fname1) > 0:
       print("You chose: %s" % fname1)
       return fname1
   else:
       print("No file chosen. Now exiting program")
       exit()

print('Please browse to the county lookup file.. File dialog window open now: ')
filename1 = browse_file2()

wb = openpyxl.load_workbook(filename)
wb1 = openpyxl.load_workbook(filename1)
sh3=wb1['COUNTY']
for sheet in wb:
    sh = wb[sheet.title]
    print('Now in sheet: ' + sheet.title)
    sh2 = wb.create_sheet("Sheet1")
    sh2.title=(sheet.title + '_Clean')
    sh2.cell(row=1, column=1).value='Name'
    sh2.cell(row=1, column=2).value='Rating'
    sh2.cell(row=1, column=3).value='Zip'
    sh2.cell(row=1, column=4).value = 'County'

    x=2
    y=1
    # iterate through excel and display data
    for i in range(2, sh.max_row + 1):
        for j in range(1, sh.max_column + 1):
            cell_obj = sh.cell(row=i, column=j)
            try:
                sub_index = cell_obj.value.find(st_cd)
            except:
                sub_index = -1
            if cell_obj.hyperlink:
                sh2.cell(row=x, column=1).value = cell_obj.value
            elif cell_obj.value == 'Rank:':
                sh2.cell(row=x, column=2).value = sh.cell(row=i+1, column=j).value.replace('/', '')
            elif sub_index != -1:
                sh2.cell(row=x, column=3).value = cell_obj.value[-5:]
                x = x+1

wb.save(filename)

df1=pd.read_excel(filename,sheet_name='Elementary_Clean')
df2=pd.read_excel(filename,sheet_name='Middle_Clean')
df3=pd.read_excel(filename,sheet_name='High_Clean')
df4=pd.read_excel(filename1,sheet_name='COUNTY')
df5=df4[df4.STATE == state_cd]
results1=df1.merge(df5,on='Zip')
results2=df2.merge(df5,on='Zip')
results3=df3.merge(df5,on='Zip')
table1 = pd.pivot_table(results1, values='Rating', index=['COUNTYNAME'], aggfunc=np.median)
table2 = pd.pivot_table(results2, values='Rating', index=['COUNTYNAME'], aggfunc=np.median)
table3 = pd.pivot_table(results3, values='Rating', index=['COUNTYNAME'], aggfunc=np.median)


with pd.ExcelWriter('Final_'+ state_cd +'.xlsx') as writer:
    results1.to_excel(writer,'Elementary_Clean',index=False)
    results2.to_excel(writer,'Middle_Clean',index=False)
    results3.to_excel(writer,'High_Clean',index=False)
    table1.to_excel(writer, 'Elementary_rating')
    table2.to_excel(writer, 'Middle_rating')
    table3.to_excel(writer, 'High_rating')



wb.save(filename)
